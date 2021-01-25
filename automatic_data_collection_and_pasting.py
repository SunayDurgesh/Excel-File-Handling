import openpyxl as xl





destination_file = input('enter the destination file name')
dest_filename = 'D:\\python\\python_projects\\excel\\'+destination_file+'.xlsx'
wb2 = xl.load_workbook(dest_filename)
ws2 = wb2.active



def sum(list):
    y=0
    for i in list:
        y+=i
    return y    

print('enter the source files')
s = list(map(str,input().split()))

x = []
for k in range(len(s)):
    filename = 'D:\\python\\python_projects\\excel\\'+s[k]+'.xlsx'
    wb = xl.load_workbook(filename)
    ws = wb.worksheets[0]


    mr = ws.max_row
    mc = ws.max_column
    x.append(mc)

    for i in range(1,mr+1):
        for j in range(1,mc+1):
            c= ws.cell(row = i,column = j)
            
            ws2.cell(row = i,column = j+(sum(x)-x[0])).value = c.value
    
wb2.save(str(dest_filename))    
    
